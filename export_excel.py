"""Utilitário para converter um JSON (no formato produzido por extraimedico.py)
em um arquivo Excel (.xlsx).

Função principal:
    json_to_excel(json_path: str, excel_path: str) -> None

Dependência: openpyxl (adicionada em requirements.txt)
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook


def _stringify(value: Any) -> str:
    """Converte valores para string legível para célula Excel.

    Regras:
      - list/tuple: itens convertidos para str e unidos por '; '
      - dict: JSON compacto
      - None: string vazia
      - demais tipos: str(value)
    """
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "; ".join(_stringify(v) for v in value if v not in (None, ""))
    if isinstance(value, dict):
        try:
            return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            return str(value)
    return str(value)


def _write_sheet_from_records(wb: Workbook, title: str, records: List[Dict[str, Any]]):
    """Cria uma planilha a partir de uma lista de dicionários homogêneos.

    Coleta o conjunto total de chaves e ordena alfabeticamente, priorizando
    algumas chaves de identificação se existirem.
    """
    if not records:
        ws = wb.create_sheet(title[:31])
        ws.append(["(vazio)"])
        return

    # Determinar colunas
    all_keys: List[str] = []
    seen = set()
    for r in records:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                all_keys.append(k)

    # Reordenar deixando campos principais primeiro caso existam
    prioridade = ["cpf", "nome", "cns", "data_nascimento"]
    ordered = [k for k in prioridade if k in all_keys] + [k for k in all_keys if k not in prioridade]

    ws = wb.create_sheet(title[:31])  # limite Excel
    ws.append(ordered)
    for r in records:
        ws.append([_stringify(r.get(col, "")) for col in ordered])

    # Congelar cabeçalho
    ws.freeze_panes = "A2"


def _write_sheet_from_simple_list(wb: Workbook, title: str, valores: List[Dict[str, Any]]):
    """Escreve lista de dicts simples (ex: identificações ausentes)."""
    _write_sheet_from_records(wb, title, valores)


def json_to_excel(json_path: str, excel_path: str) -> None:
    """Lê o arquivo JSON gerado por extraimedico.py e cria um .xlsx.

    Estrutura esperada (resumida):
    {
      "medico": {
         "blocos": [ { ..campos.. }, ...],
         "cns_nao_identificados": [ {"nome":..., "cpf":...}, ...],
         ...
      }
    }

    Serão criadas planilhas:
      - Blocos (todos os campos dos blocos)
      - CNS_Nao_Identificados
      - Mae_Nao_Identificados
      - Pai_Nao_Identificados
      - Data_Nasc_Nao_Identificados
    (somente se existirem)
    """
    json_file = Path(json_path)
    if not json_file.exists():
        raise FileNotFoundError(f"Arquivo JSON não encontrado: {json_path}")

    with json_file.open("r", encoding="utf-8") as f:
        data = json.load(f)

    medico = data.get("medico", {}) if isinstance(data, dict) else {}
    blocos = medico.get("blocos", []) if isinstance(medico, dict) else []

    wb = Workbook()
    # Remover sheet padrão criada automaticamente
    default = wb.active
    wb.remove(default)

    _write_sheet_from_records(wb, "Blocos", blocos)

    # Listas auxiliares
    mapping = {
        "cns_nao_identificados": "CNS_Nao_Identificados",
        "mae_nao_identificados": "Mae_Nao_Identificados",
        "pai_nao_identificados": "Pai_Nao_Identificados",
        "data_nascimento_nao_identificados": "Data_Nasc_Nao_Identificados",
    }
    for key, sheet_name in mapping.items():
        valores = medico.get(key, [])
        if valores:
            _write_sheet_from_simple_list(wb, sheet_name, valores)

    excel_file = Path(excel_path)
    excel_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_file))

    print(f"Excel gerado em: {excel_file}")


if __name__ == "__main__":  # uso rápido manual
    import sys
    if len(sys.argv) < 3:
        print("Uso: python export_excel.py <input_json> <output_excel>")
    else:
        json_to_excel(sys.argv[1], sys.argv[2])
